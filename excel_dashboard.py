"""
excel_dashboard.py
------------------
Creates a professional multi-sheet Excel dashboard using openpyxl.
Requires analysis_output.csv, monthly_summary.csv, volatility_summary.csv.

Sheets produced:
  1. Raw Data         - full OHLCV + indicators table with conditional formatting
  2. Monthly Summary  - per-ticker monthly KPIs
  3. Price Chart      - line chart of closing prices (all tickers)
  4. RSI Chart        - RSI trend for each ticker
  5. Volatility       - bar chart + table of annualised volatility
  6. Correlation      - colour-mapped correlation matrix
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Load Data ──────────────────────────────────────────────────────────────────
df       = pd.read_csv("analysis_output.csv")
monthly  = pd.read_csv("monthly_summary.csv")
corr     = pd.read_csv("correlation_matrix.csv", index_col=0)
vol      = pd.read_csv("volatility_summary.csv")

# ── Helpers ────────────────────────────────────────────────────────────────────
TEAL   = "1A6B72"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
LIGHT  = "F2F7F7"
DARK   = "0D3B3E"
GREEN  = "27AE60"
RED    = "E74C3C"

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(color=WHITE, bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def style_header_row(ws, row_num, fill_hex, font_color=WHITE):
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill    = header_fill(fill_hex)
            cell.font    = header_font(font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border  = thin_border()

def auto_width(ws, min_w=8, max_w=24):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

# ── Workbook ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 : Raw Data
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Raw Data")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C2"

# Title
ws1.merge_cells("A1:K1")
ws1["A1"] = "Financial Market Analysis — OHLCV + Technical Indicators"
ws1["A1"].font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
ws1["A1"].fill      = header_fill(DARK)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

COLS = ["Date","Ticker","Open","High","Low","Close","Volume",
        "Daily_Return_Pct","SMA_7","SMA_30","RSI_14",
        "BB_Upper","BB_Mid","BB_Lower","Value_Traded_Cr"]
data = df[COLS].copy()

# Header row
for ci, col_name in enumerate(COLS, start=1):
    cell = ws1.cell(row=2, column=ci, value=col_name.replace("_"," "))
style_header_row(ws1, 2, TEAL)
ws1.row_dimensions[2].height = 20

# Data rows
for ri, row in enumerate(data.itertuples(index=False), start=3):
    for ci, val in enumerate(row, start=1):
        cell = ws1.cell(row=ri, column=ci, value=val)
        cell.font   = Font(name="Calibri", size=9)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")
        if ri % 2 == 0:
            cell.fill = header_fill(LIGHT)

# Conditional formatting — Daily Return (col H = 8)
ret_col  = f"H3:H{2+len(data)}"
ws1.conditional_formatting.add(ret_col,
    CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="D5EFDF")))
ws1.conditional_formatting.add(ret_col,
    CellIsRule(operator="lessThan",    formula=["0"], fill=PatternFill("solid", fgColor="FAD7D7")))

# RSI colour scale (col K = 11)
rsi_col = f"K3:K{2+len(data)}"
ws1.conditional_formatting.add(rsi_col,
    ColorScaleRule(start_type="num", start_value=0,  start_color="E74C3C",
                   mid_type="num",   mid_value=50,   mid_color="F9E04B",
                   end_type="num",   end_value=100,  end_color="27AE60"))

auto_width(ws1)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 : Monthly Summary
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Summary")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B2"

ws2.merge_cells("A1:H1")
ws2["A1"] = "Monthly Performance Summary — All Tickers"
ws2["A1"].font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
ws2["A1"].fill      = header_fill(DARK)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for ci, col_name in enumerate(monthly.columns, start=1):
    ws2.cell(row=2, column=ci, value=col_name.replace("_"," "))
style_header_row(ws2, 2, GOLD, WHITE)

for ri, row in enumerate(monthly.itertuples(index=False), start=3):
    for ci, val in enumerate(row, start=1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.font   = Font(name="Calibri", size=9)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")
        if ri % 2 == 0:
            cell.fill = header_fill(LIGHT)

# Monthly return colour scale
mr_col = f"H3:H{2+len(monthly)}"
ws2.conditional_formatting.add(mr_col,
    ColorScaleRule(start_type="min", start_color="E74C3C",
                   mid_type="num",   mid_value=0,  mid_color="FFFFFF",
                   end_type="max",   end_color="27AE60"))
auto_width(ws2)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 : Price Chart
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Price Chart")
ws3.sheet_view.showGridLines = False

ws3["A1"] = "Closing Price Trend — All Tickers"
ws3["A1"].font = Font(name="Calibri", bold=True, size=12, color=DARK)

# Pivot: one row per date, one col per ticker
pivot = df.pivot_table(index="Date", columns="Ticker", values="Close").reset_index()
pivot_cols = ["Date"] + sorted([c for c in pivot.columns if c != "Date"])
pivot = pivot[pivot_cols]

for ci, col in enumerate(pivot_cols, start=1):
    ws3.cell(row=3, column=ci, value=col)
style_header_row(ws3, 3, TEAL)

for ri, row in enumerate(pivot.itertuples(index=False), start=4):
    for ci, val in enumerate(row, start=1):
        ws3.cell(row=ri, column=ci, value=val)

n_rows = len(pivot)
chart  = LineChart()
chart.title         = "Stock Closing Prices"
chart.style         = 10
chart.y_axis.title  = "Price (INR)"
chart.x_axis.title  = "Trading Day"
chart.height        = 14
chart.width         = 26

for i, ticker in enumerate(pivot_cols[1:], start=2):
    ref  = Reference(ws3, min_col=i, max_col=i, min_row=3, max_row=3+n_rows)
    chart.add_data(ref, titles_from_data=True)

ws3.add_chart(chart, "A5")

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 : RSI Chart
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("RSI Chart")
ws4.sheet_view.showGridLines = False
ws4["A1"] = "14-Day RSI Trend — All Tickers"
ws4["A1"].font = Font(name="Calibri", bold=True, size=12, color=DARK)

rsi_pivot = df.pivot_table(index="Date", columns="Ticker", values="RSI_14").reset_index()
rsi_cols  = ["Date"] + sorted([c for c in rsi_pivot.columns if c != "Date"])
rsi_pivot = rsi_pivot[rsi_cols]

for ci, col in enumerate(rsi_cols, start=1):
    ws4.cell(row=3, column=ci, value=col)
style_header_row(ws4, 3, GOLD)

for ri, row in enumerate(rsi_pivot.itertuples(index=False), start=4):
    for ci, val in enumerate(row, start=1):
        ws4.cell(row=ri, column=ci, value=val)

n_rsi   = len(rsi_pivot)
rsi_chart = LineChart()
rsi_chart.title        = "RSI (14-Day)"
rsi_chart.style        = 10
rsi_chart.y_axis.title = "RSI Value"
rsi_chart.y_axis.scaling.min = 0
rsi_chart.y_axis.scaling.max = 100
rsi_chart.height = 14
rsi_chart.width  = 26

for i, ticker in enumerate(rsi_cols[1:], start=2):
    ref = Reference(ws4, min_col=i, max_col=i, min_row=3, max_row=3+n_rsi)
    rsi_chart.add_data(ref, titles_from_data=True)

ws4.add_chart(rsi_chart, "A5")

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 : Volatility
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Volatility")
ws5.sheet_view.showGridLines = False
ws5["A1"] = "Annualised Volatility by Ticker"
ws5["A1"].font = Font(name="Calibri", bold=True, size=12, color=DARK)

for ci, col in enumerate(vol.columns, start=1):
    ws5.cell(row=3, column=ci, value=col)
style_header_row(ws5, 3, TEAL)

for ri, row in enumerate(vol.itertuples(index=False), start=4):
    for ci, val in enumerate(row, start=1):
        ws5.cell(row=ri, column=ci, value=val)

bar = BarChart()
bar.type   = "col"
bar.title  = "Annualised Volatility (%)"
bar.style  = 10
bar.y_axis.title = "Volatility"
bar.x_axis.title = "Ticker"
bar.height = 12
bar.width  = 18

data_ref   = Reference(ws5, min_col=2, max_col=2, min_row=3, max_row=3+len(vol))
cats_ref   = Reference(ws5, min_col=1, max_col=1, min_row=4, max_row=3+len(vol))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
ws5.add_chart(bar, "D3")
auto_width(ws5)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 : Correlation Matrix
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Correlation")
ws6.sheet_view.showGridLines = False
ws6["A1"] = "Close Price Correlation Matrix"
ws6["A1"].font = Font(name="Calibri", bold=True, size=12, color=DARK)

tickers = corr.index.tolist()
# Header row
for ci, t in enumerate(tickers, start=2):
    ws6.cell(row=3, column=ci, value=t)
style_header_row(ws6, 3, DARK)

for ri, row_ticker in enumerate(tickers, start=4):
    ws6.cell(ri, 1, row_ticker).font = Font(bold=True)
    for ci, col_ticker in enumerate(tickers, start=2):
        val  = corr.loc[row_ticker, col_ticker]
        cell = ws6.cell(ri, ci, round(float(val), 4))
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_border()
        cell.font      = Font(name="Calibri", size=9)

corr_range = f"B4:F{3+len(tickers)}"
ws6.conditional_formatting.add(corr_range,
    ColorScaleRule(start_type="num", start_value=-1, start_color="E74C3C",
                   mid_type="num",   mid_value=0,    mid_color="FFFFFF",
                   end_type="num",   end_value=1,    end_color="1A6B72"))
auto_width(ws6)

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save("financial_market_dashboard.xlsx")
print("financial_market_dashboard.xlsx saved — 6 sheets")
